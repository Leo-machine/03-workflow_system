"""表格导入流程草稿：只新建 draft，不覆盖已有流程，不静默建人/团队。"""
from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field

from fastapi import HTTPException
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app.models import BusinessDomain, Flow, Person, Unit, User
from app.schemas import FlowImportFlowPlan, FlowImportIssue, FlowImportResult, GuideItemIn, StepDefinitionIn

HEADERS = [
    "业务域",
    "流程名称",
    "流程说明",
    "环节编号",
    "环节名称",
    "环节任务",
    "系统名",
    "动作说明",
    "系统链接",
    "依据注意",
    "责任团队",
    "责任人",
    "直接领导",
]

MAX_ROWS = 2000
MAX_FLOWS = 50
NAME_SPLIT = re.compile(r"[,，;；、]+")


@dataclass
class ParsedGuide:
    row: int
    system_name: str
    action_text: str
    url: str | None
    note: str | None
    unit_name: str
    person_names: list[str]
    escalation_name: str


@dataclass
class ParsedStep:
    code: str
    name: str
    task: str
    guides: list[ParsedGuide] = field(default_factory=list)


@dataclass
class ParsedFlow:
    domain_name: str
    name: str
    description: str
    first_row: int
    steps: list[ParsedStep] = field(default_factory=list)


def template_csv() -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(HEADERS)
    writer.writerow([
        "#填写说明：以下为算力卡生命周期样例，请将业务域、责任团队、责任人与直接领导替换为系统台账中的实际名称；可删除本说明行",
    ])
    sample = [
        ["配件管理", "算力卡全生命周期溯源（导入示例）", "算力卡从到货入库到归还入库的全过程操作指引", "S01", "到货入库", "核对到货资料并建立配件资产档案", "配件管理系统", "按到货单录入固定资产编号、序列号、供应商、合同、单价、产权单位和维保到期日", "https://example.internal/accessories/inbound", "核对实物铭牌与到货单信息一致", "", "", ""],
        ["配件管理", "算力卡全生命周期溯源（导入示例）", "", "S02", "装机", "将库存算力卡安装到目标服务器", "配件管理系统", "选择目标服务器并建立配件与服务器安装关系，将状态由在库变更为在用", "https://example.internal/accessories/install", "装机前确认服务器处于可操作状态", "", "", ""],
        ["配件管理", "算力卡全生命周期溯源（导入示例）", "", "S03", "拆下", "服务器下线检修后拆除算力卡", "配件管理系统", "确认服务器已切换为未投运，解除安装关系并将算力卡状态变更为在库", "https://example.internal/accessories/remove", "禁止在服务器运行状态下直接拆除", "", "", ""],
        ["配件管理", "算力卡全生命周期溯源（导入示例）", "", "S04", "借出", "审批通过后向兄弟单位调出", "配件管理系统", "创建借出单，登记借用单位、经办人、预计归还时间和关联工单", "https://example.internal/accessories/lend", "审批通过后方可出库", "", "", ""],
        ["配件管理", "算力卡全生命周期溯源（导入示例）", "", "S05", "归还", "验收归还配件并完成入库", "配件管理系统", "核对序列号和配件状态，关闭借出记录并重新入库", "https://example.internal/accessories/return", "如有损坏需同步登记异常", "", "", ""],
    ]
    writer.writerows(sample)
    return ("\ufeff" + buf.getvalue()).encode("utf-8")


def _decode_text(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("文件编码无法识别，请另存为 CSV UTF-8 或 Excel .xlsx")


def _cells_from_csv(data: bytes) -> list[list[str]]:
    text = _decode_text(data)
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
    except csv.Error:
        dialect = csv.excel
    rows: list[list[str]] = []
    for row in csv.reader(io.StringIO(text), dialect):
        rows.append([cell.strip() for cell in row])
    return rows


def _cells_from_xlsx(data: bytes) -> list[list[str]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(["" if cell is None else str(cell).strip() for cell in row])
    return rows


def load_table_rows(filename: str, data: bytes) -> list[list[str]]:
    name = (filename or "").lower()
    if name.endswith(".xlsx") or data[:2] == b"PK":
        return _cells_from_xlsx(data)
    return _cells_from_csv(data)


def _split_names(raw: str) -> list[str]:
    if not raw:
        return []
    names = [part.strip() for part in NAME_SPLIT.split(raw) if part.strip()]
    seen: set[str] = set()
    unique: list[str] = []
    for name in names:
        if name in seen:
            continue
        seen.add(name)
        unique.append(name)
    return unique


HEADER_ALIASES = {"升级联系人": "直接领导"}


def _header_map(header: list[str]) -> dict[str, int] | str:
    index: dict[str, int] = {}
    for i, name in enumerate(header):
        if not name:
            continue
        index[HEADER_ALIASES.get(name, name)] = i
    missing = [col for col in HEADERS if col not in index]
    if missing:
        return f"表头缺少列：{'、'.join(missing)}"
    return index


def parse_flows_from_rows(rows: list[list[str]]) -> tuple[list[ParsedFlow], list[FlowImportIssue]]:
    issues: list[FlowImportIssue] = []
    if not rows:
        return [], [FlowImportIssue(row=0, message="文件为空")]

    header_idx = 0
    while header_idx < len(rows) and not any(rows[header_idx]):
        header_idx += 1
    if header_idx >= len(rows):
        return [], [FlowImportIssue(row=0, message="文件为空")]

    mapped = _header_map(rows[header_idx])
    if isinstance(mapped, str):
        return [], [FlowImportIssue(row=header_idx + 1, message=mapped)]

    flows: list[ParsedFlow] = []
    by_key: dict[tuple[str, str], ParsedFlow] = {}
    data_rows = 0
    for offset, raw in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        if not any(raw):
            continue
        if raw[0].startswith("#"):
            continue
        data_rows += 1
        if data_rows > MAX_ROWS:
            issues.append(FlowImportIssue(row=offset, message=f"超过 {MAX_ROWS} 行上限"))
            break

        def col(name: str) -> str:
            idx = mapped[name]
            return raw[idx].strip() if idx < len(raw) else ""

        domain_name = col("业务域")
        flow_name = col("流程名称")
        code = col("环节编号")
        step_name = col("环节名称")
        if not domain_name or not flow_name or not code or not step_name:
            issues.append(
                FlowImportIssue(row=offset, message="业务域、流程名称、环节编号、环节名称不能为空")
            )
            continue

        key = (domain_name, flow_name)
        parsed = by_key.get(key)
        if parsed is None:
            if len(by_key) >= MAX_FLOWS:
                issues.append(FlowImportIssue(row=offset, message=f"单次最多导入 {MAX_FLOWS} 条流程"))
                continue
            parsed = ParsedFlow(
                domain_name=domain_name,
                name=flow_name,
                description=col("流程说明"),
                first_row=offset,
            )
            by_key[key] = parsed
            flows.append(parsed)
        elif col("流程说明") and not parsed.description:
            parsed.description = col("流程说明")

        step = next((s for s in parsed.steps if s.code == code), None)
        if step is None:
            step = ParsedStep(code=code, name=step_name, task=col("环节任务"))
            parsed.steps.append(step)
        else:
            if step.name != step_name:
                issues.append(
                    FlowImportIssue(
                        row=offset,
                        message=f"环节编号 {code} 的名称与第 {parsed.first_row} 行不一致",
                    )
                )
            if col("环节任务") and not step.task:
                step.task = col("环节任务")

        system_name = col("系统名")
        action_text = col("动作说明")
        if not system_name and not action_text and not col("责任人") and not col("责任团队"):
            continue
        if not system_name or not action_text:
            issues.append(FlowImportIssue(row=offset, message="系统名与动作说明不能为空"))
            continue
        url = col("系统链接") or None
        if url and not (url.startswith("http://") or url.startswith("https://")):
            issues.append(FlowImportIssue(row=offset, message="系统链接仅支持 http/https"))
        step.guides.append(
            ParsedGuide(
                row=offset,
                system_name=system_name,
                action_text=action_text,
                url=url,
                note=col("依据注意") or None,
                unit_name=col("责任团队"),
                person_names=_split_names(col("责任人")),
                escalation_name=col("直接领导"),
            )
        )

    if data_rows == 0 and not issues:
        issues.append(FlowImportIssue(row=header_idx + 1, message="没有可导入的数据行"))
    return flows, issues


def _resolve_unit(name: str, units: dict[str, Unit], row: int, issues: list[FlowImportIssue]) -> Unit | None:
    unit = units.get(name)
    if unit is None:
        issues.append(FlowImportIssue(row=row, message=f"责任团队「{name}」不在台账中"))
    return unit


def _resolve_person(
    name: str,
    *,
    unit: Unit | None,
    persons: list[Person],
    row: int,
    issues: list[FlowImportIssue],
    label: str,
) -> Person | None:
    matches = [p for p in persons if p.name == name]
    if unit is not None:
        in_unit = [p for p in matches if p.unit_id == unit.id]
        if len(in_unit) == 1:
            return in_unit[0]
        if len(in_unit) > 1:
            issues.append(FlowImportIssue(row=row, message=f"{label}「{name}」在该团队内不唯一"))
            return None
        if matches:
            issues.append(FlowImportIssue(row=row, message=f"{label}「{name}」不属于团队「{unit.name}」"))
            return None
        issues.append(FlowImportIssue(row=row, message=f"{label}「{name}」不在人员台账中"))
        return None
    if len(matches) == 1:
        return matches[0]
    if len(matches) > 1:
        issues.append(FlowImportIssue(row=row, message=f"{label}「{name}」在台账中不唯一，请填写所属团队"))
        return None
    issues.append(FlowImportIssue(row=row, message=f"{label}「{name}」不在人员台账中"))
    return None


def build_definitions(
    db: Session, flows: list[ParsedFlow], issues: list[FlowImportIssue]
) -> list[tuple[ParsedFlow, BusinessDomain, list[StepDefinitionIn]]]:
    domains = {d.name: d for d in db.scalars(select(BusinessDomain)).all()}
    units = {u.name: u for u in db.scalars(select(Unit)).all()}
    persons = list(db.scalars(select(Person).options(selectinload(Person.unit))).all())
    existing_names: dict[int, set[str]] = {}
    planned_names: dict[int, set[str]] = {}
    result: list[tuple[ParsedFlow, BusinessDomain, list[StepDefinitionIn]]] = []

    for parsed in flows:
        domain = domains.get(parsed.domain_name)
        if domain is None:
            issues.append(
                FlowImportIssue(
                    row=parsed.first_row,
                    message=f"业务域「{parsed.domain_name}」不存在，请先在系统中创建",
                )
            )
            continue
        names = existing_names.setdefault(
            domain.id,
            set(db.scalars(select(Flow.name).where(Flow.domain_id == domain.id)).all()),
        )
        planned = planned_names.setdefault(domain.id, set())
        if parsed.name in names or parsed.name in planned:
            issues.append(
                FlowImportIssue(
                    row=parsed.first_row,
                    message=f"业务域「{domain.name}」已有流程「{parsed.name}」，导入只新建草稿、不覆盖已有流程",
                )
            )
            continue
        planned.add(parsed.name)

        steps_in: list[StepDefinitionIn] = []
        for step in parsed.steps:
            guides_in: list[GuideItemIn] = []
            for guide in step.guides:
                unit = None
                if guide.unit_name:
                    unit = _resolve_unit(guide.unit_name, units, guide.row, issues)
                elif guide.person_names:
                    issues.append(FlowImportIssue(row=guide.row, message="请先填写责任团队再填写责任人"))
                person_ids: list[int] = []
                for name in guide.person_names:
                    person = _resolve_person(
                        name, unit=unit, persons=persons, row=guide.row, issues=issues, label="责任人"
                    )
                    if person is not None:
                        person_ids.append(person.id)
                escalation_id = None
                if guide.escalation_name:
                    esc = _resolve_person(
                        guide.escalation_name,
                        unit=None,
                        persons=persons,
                        row=guide.row,
                        issues=issues,
                        label="直接领导",
                    )
                    if esc is not None:
                        escalation_id = esc.id
                guides_in.append(
                    GuideItemIn(
                        system_name=guide.system_name,
                        action_text=guide.action_text,
                        url=guide.url,
                        note=guide.note,
                        unit_id=unit.id if unit else None,
                        person_ids=person_ids,
                        escalation_person_id=escalation_id,
                    )
                )
            steps_in.append(
                StepDefinitionIn(code=step.code, name=step.name, task=step.task, guide=guides_in)
            )
        result.append((parsed, domain, steps_in))
    return result


def preview_or_commit(
    db: Session,
    *,
    filename: str,
    data: bytes,
    admin: User,
    commit: bool,
) -> FlowImportResult:
    from app.routers.flows import (
        _add_flow_log,
        _insert_definition_steps,
        _next_order_index,
        _unique_slug,
        _validate_definition_steps,
        _load_definition_persons,
    )

    issues: list[FlowImportIssue] = []
    try:
        rows = load_table_rows(filename, data)
    except ValueError as exc:
        return FlowImportResult(
            ok=False, committed=False, issues=[FlowImportIssue(row=0, message=str(exc))]
        )
    except Exception:
        return FlowImportResult(
            ok=False,
            committed=False,
            issues=[FlowImportIssue(row=0, message="无法读取表格，请使用模板 CSV 或 .xlsx")],
        )

    parsed_flows, parse_issues = parse_flows_from_rows(rows)
    issues.extend(parse_issues)
    built = build_definitions(db, parsed_flows, issues)

    created_ids: list[int] = []
    plans: list[FlowImportFlowPlan] = []
    for parsed, domain, steps_in in built:
        plans.append(
            FlowImportFlowPlan(
                domain_name=domain.name,
                flow_name=parsed.name,
                step_count=len(steps_in),
                guide_count=sum(len(step.guide) for step in steps_in),
            )
        )
        try:
            persons_by_id = _load_definition_persons(db, steps_in)
            _validate_definition_steps(steps_in, persons_by_id)
        except HTTPException as exc:
            issues.append(FlowImportIssue(row=parsed.first_row, message=str(exc.detail)))

    if issues:
        return FlowImportResult(ok=False, committed=False, issues=issues, flows=plans)

    if not commit:
        return FlowImportResult(ok=True, committed=False, issues=[], flows=plans)

    for parsed, domain, steps_in in built:
        flow = Flow(
            slug=_unique_slug(db),
            domain_id=domain.id,
            name=parsed.name,
            description=parsed.description,
            status="draft",
            order_index=_next_order_index(db, domain.id),
            updated_by=admin.username,
        )
        db.add(flow)
        db.flush()
        _insert_definition_steps(db, flow, steps_in)
        _add_flow_log(
            db,
            flow=flow,
            field="import",
            old_value=None,
            new_value="draft",
            admin=admin,
            new_name=flow.name,
        )
        created_ids.append(flow.id)

    db.commit()
    return FlowImportResult(ok=True, committed=True, issues=[], flows=plans, created_flow_ids=created_ids)
